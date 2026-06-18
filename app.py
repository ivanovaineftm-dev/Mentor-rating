from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from flask import Flask, flash, redirect, render_template, request, send_file, url_for
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from werkzeug.utils import secure_filename

SCORE_COLUMN_INDEX = 14  # Column N, immediately after M.
DETAIL_COLUMN_INDEX = SCORE_COLUMN_INDEX + 1  # Column O, immediately after score.
SCORE_HEADER = "Балл"
DETAIL_HEADER = "Подробный расчет"

app = Flask(__name__)
app.config["SECRET_KEY"] = "mentor-rating-local-secret"
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def is_numeric(value: Any) -> bool:
    if is_blank(value) or isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        normalized = value.strip().replace(",", ".")
        try:
            float(normalized)
        except ValueError:
            return False
        return True
    return False


def is_passed(value: Any) -> bool:
    return isinstance(value, str) and value.strip().casefold() == "сдан"


def get_score_components(column_b: Any, column_m: Any, column_e: Any, column_h: Any) -> dict[str, float]:
    b_coef = 1 if is_blank(column_b) else 0 if is_numeric(column_b) else 1
    m_coef = 1 if is_blank(column_m) else 0 if is_numeric(column_m) else 0
    e_coef = 1 if is_passed(column_e) else 0
    h_coef = 1 if is_passed(column_h) else 0
    base_score = (0.4 * b_coef) + (0.35 * m_coef) + (0.25 * e_coef)
    h_multiplier = 1 + (0.2 * h_coef)
    score = round(base_score * h_multiplier, 4)

    return {
        "b_coef": b_coef,
        "m_coef": m_coef,
        "e_coef": e_coef,
        "h_coef": h_coef,
        "base_score": round(base_score, 4),
        "h_multiplier": round(h_multiplier, 4),
        "score": score,
    }


def calculate_score(column_b: Any, column_m: Any, column_e: Any, column_h: Any) -> float:
    return get_score_components(column_b, column_m, column_e, column_h)["score"]


def format_score_calculation(column_b: Any, column_m: Any, column_e: Any, column_h: Any) -> str:
    components = get_score_components(column_b, column_m, column_e, column_h)
    return (
        f"B={components['b_coef']}: 0,4×{components['b_coef']}={0.4 * components['b_coef']:.2f}; "
        f"M={components['m_coef']}: 0,35×{components['m_coef']}={0.35 * components['m_coef']:.2f}; "
        f"E={components['e_coef']}: 0,25×{components['e_coef']}={0.25 * components['e_coef']:.2f}; "
        f"база={components['base_score']:.2f}; "
        f"H={components['h_coef']}: множитель 1+0,2×{components['h_coef']}={components['h_multiplier']:.2f}; "
        f"итог={components['base_score']:.2f}×{components['h_multiplier']:.2f}={components['score']:.4f}"
    )


def remove_existing_result_columns(sheet: Worksheet) -> None:
    for column_index, header_text in (
        (DETAIL_COLUMN_INDEX, DETAIL_HEADER),
        (SCORE_COLUMN_INDEX, SCORE_HEADER),
    ):
        if sheet.max_column >= column_index:
            header = sheet.cell(row=1, column=column_index).value
            if isinstance(header, str) and header.strip().casefold() == header_text.casefold():
                sheet.delete_cols(column_index, 1)


def process_worksheet(sheet: Worksheet) -> None:
    max_data_row = sheet.max_row
    remove_existing_result_columns(sheet)
    sheet.insert_cols(SCORE_COLUMN_INDEX, 2)
    sheet.cell(row=1, column=SCORE_COLUMN_INDEX, value=SCORE_HEADER)
    sheet.cell(row=1, column=DETAIL_COLUMN_INDEX, value=DETAIL_HEADER)

    for row in range(2, max_data_row + 1):
        column_b = sheet.cell(row=row, column=2).value
        column_m = sheet.cell(row=row, column=13).value
        column_e = sheet.cell(row=row, column=5).value
        column_h = sheet.cell(row=row, column=8).value
        score = calculate_score(column_b, column_m, column_e, column_h)
        detail = format_score_calculation(column_b, column_m, column_e, column_h)
        sheet.cell(row=row, column=SCORE_COLUMN_INDEX, value=score)
        sheet.cell(row=row, column=DETAIL_COLUMN_INDEX, value=detail)


def process_workbook(file_stream: BytesIO) -> BytesIO:
    workbook = load_workbook(file_stream)
    for sheet in workbook.worksheets:
        process_worksheet(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/calculate", methods=["POST"])
def calculate():
    uploaded_file = request.files.get("file")
    if uploaded_file is None or uploaded_file.filename == "":
        flash("Выберите Excel-файл для обработки.", "error")
        return redirect(url_for("index"))

    filename = secure_filename(uploaded_file.filename) or "rating"

    try:
        output = process_workbook(BytesIO(uploaded_file.read()))
    except Exception as exc:  # noqa: BLE001 - user-facing upload validation boundary.
        flash(f"Не удалось обработать файл: {exc}", "error")
        return redirect(url_for("index"))

    stem = Path(filename).stem or "rating"
    download_name = f"{stem}_with_scores.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=download_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
