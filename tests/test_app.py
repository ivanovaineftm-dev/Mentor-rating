from io import BytesIO
from math import nan
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pytest

openpyxl = pytest.importorskip("openpyxl")
pytest.importorskip("flask")

from openpyxl import Workbook, load_workbook

from app import calculate_score, has_value, is_blank, normalized_text, process_workbook


def test_blank_detection_matches_rules():
    assert is_blank(None)
    assert is_blank("")
    assert is_blank("   ")
    assert is_blank(nan)
    assert not is_blank("text")
    assert not is_blank(0)


def test_has_value_treats_any_non_blank_type_as_filled():
    assert has_value(123)
    assert has_value("text")
    assert has_value("0")
    assert has_value(False)
    assert not has_value(None)
    assert not has_value("   ")


def test_normalized_text_trims_and_lowercases_strings_only():
    assert normalized_text(" ПРОЙДЕН ") == "пройден"
    assert normalized_text(" СДАН ") == "сдан"
    assert normalized_text(123) == ""


def test_calculate_score_user_example_filled_b_blank_m_failed_e_passed_h():
    assert calculate_score("любое значение", None, "не пройден", "сдан") == 0.2


def test_calculate_score_user_example_blank_b_blank_m_failed_e_passed_h():
    assert calculate_score(None, None, "не пройден", "сдан") == 0.6


def test_calculate_score_uses_exact_allowed_e_and_h_values():
    assert calculate_score(None, "перевод", " ПрОйДеН ", " СдАн ") == 1.0
    assert calculate_score(None, "перевод", "сдан", "пройден") == 0.6
    assert calculate_score(None, "перевод", "другое", "другое") == 0.6


def test_calculate_score_b_and_m_do_not_require_numbers():
    assert calculate_score("уволен", "перевод", "пройден", "сдан") == 0.6


def test_calculate_score_blank_m_is_zero_and_blank_b_is_one():
    assert calculate_score(None, None, "не пройден", "не сдан") == 0.4


def test_process_workbook_inserts_score_after_m_and_preserves_following_columns():
    workbook = Workbook()
    sheet = workbook.active
    headers = [f"H{i}" for i in range(1, 16)]
    for column, value in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=value)

    sheet.cell(row=2, column=2, value=None)
    sheet.cell(row=2, column=5, value="пройден")
    sheet.cell(row=2, column=8, value="сдан")
    sheet.cell(row=2, column=13, value=100)
    sheet.cell(row=2, column=14, value="original N")

    source = BytesIO()
    workbook.save(source)
    source.seek(0)

    result = process_workbook(source)
    processed = load_workbook(result)
    processed_sheet = processed.active

    assert processed_sheet.cell(row=1, column=14).value == "Балл"
    assert processed_sheet.cell(row=2, column=14).value == 1.0
    assert processed_sheet.cell(row=1, column=15).value == "H14"
    assert processed_sheet.cell(row=2, column=15).value == "original N"


def test_process_workbook_replaces_existing_score_column():
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=14, value="Балл")
    sheet.cell(row=2, column=14, value=999)

    source = BytesIO()
    workbook.save(source)
    source.seek(0)

    result = process_workbook(source)
    processed = load_workbook(result)

    assert processed.active.cell(row=1, column=14).value == "Балл"
    assert processed.active.cell(row=2, column=14).value == 0.4


def test_calculate_route_accepts_xlsx_content():
    workbook = Workbook()
    sheet = workbook.active
    sheet.cell(row=2, column=5, value="пройден")

    source = BytesIO()
    workbook.save(source)
    source.seek(0)

    from app import app

    app.config.update(TESTING=True)
    with app.test_client() as client:
        response = client.post(
            "/calculate",
            data={"file": (source, "rating.xlsx")},
            content_type="multipart/form-data",
        )

    assert response.status_code == 200
    assert response.headers["Content-Disposition"].startswith(
        "attachment; filename=rating_with_scores.xlsx"
    )

    processed = load_workbook(BytesIO(response.data))
    assert processed.active.cell(row=1, column=14).value == "Балл"
    assert processed.active.cell(row=2, column=14).value == 0.6
